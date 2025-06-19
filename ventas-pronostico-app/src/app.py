import streamlit as st
import pandas as pd

from utils.excel_importer import import_excel
from utils.charts import mostrar_grafico_clustered_line
from models.forecasting import ForecastingModel
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

st.set_page_config(layout="wide")
st.markdown("""
    <style>
        .stDataFrame, .stTable {
            width: 100% !important;
        }
    </style>
""", unsafe_allow_html=True)

def main():
    st.title("WIREBI - Aplicación de Pronóstico de Ventas")

    uploaded_file = st.file_uploader("Sube un archivo Excel con datos de clientes, productos y cantidades", type=["xlsx"])
    
    if uploaded_file is not None:
        data = import_excel(uploaded_file)
        data['Month'] = pd.to_datetime(data['Month'])
        data_pivot = data.pivot_table(
            index=['DPG', 'SKU'],
            columns=data['Month'].dt.strftime('%Y-%m'),
            values='Sum of Quantity',
            aggfunc='sum'
        ).reset_index()

        # Inicializar ajustes y valores ajustados en session_state
        if 'ajustes' not in st.session_state:
            st.session_state['ajustes'] = data_pivot.copy()
            for col in data_pivot.columns:
                if col not in ['DPG', 'SKU']:
                    st.session_state['ajustes'][col] = 0
        if 'data_ajustada' not in st.session_state:
            st.session_state['data_ajustada'] = data_pivot.copy()

        # Mostrar primero los valores ajustados
        st.write("Valores ajustados (original + ajuste):")
        st.dataframe(st.session_state['data_ajustada'])

        ajustes_aplicados = False
        if 'ajustes' in st.session_state:
            # Solo columnas de meses
            cols_ajuste = [col for col in st.session_state['ajustes'].columns if col not in ['DPG', 'SKU']]
            ajustes_aplicados = (st.session_state['ajustes'][cols_ajuste].to_numpy() != 0).any()

        # Mostrar gráfico de líneas agrupadas
        if ajustes_aplicados:
            mostrar_grafico_clustered_line(data, st.session_state['data_ajustada'], st.session_state.get('forecast'))
        else:
            mostrar_grafico_clustered_line(data, None, st.session_state.get('forecast'))
       
        # Formulario para editar y aplicar ajustes
        with st.form("ajustes_form"):
            st.write("Ajustes (edita aquí para sumar/restar cantidades):")
            ajustes_editados = st.data_editor(
                st.session_state['ajustes'],
                key="ajustes_editor"
            )
            aplicar = st.form_submit_button("Aplicar ajustes")

        if aplicar:
            # Actualiza primero el session_state con los datos editados
            st.session_state['ajustes'] = ajustes_editados
            # Sumar los ajustes a los valores originales (solo columnas de meses)
            data_ajustada = data_pivot.copy()
            for col in data_pivot.columns:
                if col not in ['DPG', 'SKU']:
                    data_ajustada[col] = data_pivot[col] + st.session_state['ajustes'][col]
            st.session_state['data_ajustada'] = data_ajustada

            st.warning("Si acabas de editar una celda, por favor presiona 'Aplicar ajustes' una segunda vez para que se apliquen los cambios.")



        n_months = st.number_input(
            "¿Cuántos meses quieres pronosticar?", min_value=1, max_value=48, value=12
        )

        data_ajustada_long = st.session_state['data_ajustada'].melt(
            id_vars=['DPG', 'SKU'],
            var_name='Month',
            value_name='Sum of Quantity'
        )
        data_ajustada_long = data_ajustada_long[~data_ajustada_long['Month'].isin(['index'])]
        data_ajustada_long['Month'] = pd.to_datetime(data_ajustada_long['Month'], format='%Y-%m')

        model = ForecastingModel(data_ajustada_long)
        model.train_model()
        
        if st.button("Realizar Pronóstico"):
            last_date = data_ajustada_long['Month'].max()
            future_dates = pd.date_range(
                start=last_date + pd.offsets.MonthBegin(1),
                periods=n_months,
                freq='MS'
            )
            future_data = pd.DataFrame({'Month': future_dates})

            predictions = model.predict_sales(future_data)
            pred_pivot = predictions.pivot_table(
                index=['DPG', 'SKU'],
                columns=predictions['Month'].dt.strftime('%Y-%m'),
                values='Forecast',
                aggfunc='sum'
            ).reset_index()
            st.write("Pronósticos de ventas ajustados:")
            st.dataframe(pred_pivot)
            # Guardar en formato largo para el gráfico
            pred_long = pred_pivot.melt(
                id_vars=['DPG', 'SKU'],
                var_name='Month',
                value_name='Valor'
            )
            pred_long['Month'] = pd.to_datetime(pred_long['Month'], format='%Y-%m')
            st.session_state['forecast'] = pred_long

        # Botón único para generar y descargar el pronóstico en Excel
        if 'forecast' in st.session_state:
            output_final = None
            if st.button("Generar el pronóstico en formato Excel"):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    st.session_state['forecast'].to_excel(writer, index=False, sheet_name='Pronostico')
                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active
                max_row = ws.max_row
                max_col = ws.max_column
                last_col_letter = get_column_letter(max_col)
                table_ref = f"A1:{last_col_letter}{max_row}"
                table = Table(displayName="Pronostico", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                ws.add_table(table)
                output_final = BytesIO()
                wb.save(output_final)
                output_final.seek(0)
                st.download_button(
                    label="Descargar pronóstico (XLSX)",
                    data=output_final,
                    file_name="pronostico_forecast.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

if __name__ == "__main__":
    main()